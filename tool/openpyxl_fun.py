from openpyxl import load_workbook
from tool.eduMySQL import edu_pyMySql


class Excel_data(object):
    def __init__(self, path, page):
        self.local = path
        self.requirename = page

    # A为Excel的行号
    def excel_isTrue(self, casename):
        # 获取Excel文件
        self.wb = load_workbook(self.local)
        # 获取Excel子页
        self.ws = self.wb[self.requirename]
        self.a = len(self.ws['A']) + 1
        for i in range(1, len(self.ws['A'])):
            # 循环判断A行测试数据与代码里的测试用例函数名相同的
            if self.ws['A'][i].value == casename:
                # print(self.ws['A'][i].value)
                self.a = i
            else:
                continue
        isTrue = self.ws.cell(row=self.ws['A'][self.a].row, column=10).value
        # print(isTrue)
        return isTrue

    # 获取预期结果
    def expected_result(self):
        # 判断A的i行这一列第8格判断用例是获取预期结果的方法，SQL是执行数据库获取动态结果，assert是直接获取文本
        var = self.ws.cell(row=self.ws['A'][self.a].row, column=8).value
        if var == 'sql':
            sql = self.ws.cell(row=self.ws['A'][self.a].row, column=9).value
            if sql is not None:
                reality = edu_pyMySql(sql)
                return reality
            else:
                print('这里没有SQL语句')
        elif var == 'assert':
            Expect = self.ws.cell(row=self.ws['A'][self.a].row, column=9).value
            return Expect

    # 测试数据
    def test_data(self):
        result = self.ws.cell(row=self.ws['A'][self.a].row, column=7).value
        if result is not None:
            clom = self.ws.cell(row=self.ws['A'][self.a].row, column=6).value
            stuff = result.find(',')
            key = clom.split(',')
            # 判断是否数据
            if stuff != -1:
                values = result.split(',')
                data = {}
                for j in range(len(key)):
                    data[key[j]] = values[j]

                return data

    # 测试路径
    def host_url(self):
        host = self.ws.cell(row=self.ws['A'][self.a].row, column=3).value
        if host is not None:

            apipath = self.ws.cell(row=self.ws['A'][self.a].row, column=4).value
            if apipath is not None:
                url = host + apipath
                # print(url)
                return url

    def excel_data(self, casename):
        excel_isTrue = self.excel_isTrue(casename)
        if excel_isTrue == 'yes':
            url = self.host_url()
            data = self.test_data()
            expected_result = self.expected_result()
            return url, data, expected_result
        elif excel_isTrue == 'no':
            pass

    def excel_close(self):
        self.wb.close()


if __name__ == '__main__':
    path = '../file_generalList/Excel_TestSet/edu_login_v01.xlsx'
    page = 'addTeacher'
    casename = 'test_add_teacher_1'
    run = Excel_data(path, page)
    a = run.excel_data(casename)
    url1 = a[0]
    data1 = a[1]
    expected_result1 = a[2]
    print(type(url1), url1)
    print(type(data1), data1)
    print(type(expected_result1), expected_result1)
